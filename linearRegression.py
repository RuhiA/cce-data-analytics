import math
import openpyxl
import sys
#@author : Manisha Nibhwani
#This program reads columns from excel and fill the values in excel

class Correlation:
    value=0
    isSignificant=0

class LinearRegression:
    m=0 
    c =0

def main() :
    global sheet
    global wb
    if(len(sys.argv)>3) :
        global filePath 
        filePath = sys.argv[1]
        wb = openpyxl.load_workbook(sys.argv[1])
        sheet = wb.active
        mapColumnsWithIndex()
        if (correlation(sys.argv[2],sys.argv[3]).isSignificant == 1):
            res = linearRegression(sys.argv[2], sys.argv[3])
            print(res.m +" "+res.c)

def mapColumnsWithIndex() :
    global columnArray
    columnArray = {}
    for i in range(sheet.max_column): 
        columnArray[i] = sheet.cell(1, i+1).value

def getIndexFromMap(key) :
    for i in range(len(columnArray)) :
        if(key == columnArray[i]) :
            return i
    return -1     


def sumOfColumn(columnName) :
    index = getIndexFromMap(columnName)
    numRows = sheet.max_row
    sumOfCol=0
    for i in range(numRows) :
        if(i+1 < numRows) :
            sumOfCol = sumOfCol + sheet.cell(i+2, index+1).value
    return sumOfCol

def meanOfColumn(columnName) :
    sum = sumOfColumn(columnName)
    return sum/(sheet.max_row-1)

def addColumnMinusMean(columnName):
    index = getIndexFromMap(columnName)
    mean = meanOfColumn(columnName)
    maxColumns = sheet.max_column
    newIndex = getIndexFromMap(columnName+"MinusMean")
    if (newIndex == -1) :
        sheet.cell(1,maxColumns+1).value = columnName+"MinusMean"
    columnArray[maxColumns] = columnName+"MinusMean"
    sumOfNewCell = 0
    for i in range(sheet.max_row) :
        if (i+1 < sheet.max_row) :
            value = sheet.cell(i+2,index + 1).value - mean
            if(newIndex == -1):
                sheet.cell(i+2,maxColumns+1).value =  value
            sumOfNewCell = sumOfNewCell + value
    wb.save(filePath)
    return sumOfNewCell

def addSquareOfColumn(columnName) :
    index = getIndexFromMap(columnName)
    maxColumns = sheet.max_column
    newIndex = getIndexFromMap(columnName+"Square")
    if (newIndex == -1) :
        sheet.cell(1,maxColumns+1).value = columnName+"Square"
    columnArray[maxColumns] = columnName+"Square"
    sumOfNewCell = 0
    for i in range(sheet.max_row) :
        if (i+1 < sheet.max_row) :
            value = sheet.cell(i+2,index + 1).value    
            if(newIndex == -1):
                sheet.cell(i+2,maxColumns+1).value =  value * value
            sumOfNewCell = sumOfNewCell + (value * value)
    wb.save(filePath)
    return sumOfNewCell

def addProductOfColumns(column1,column2, newColumnName) :
    index1 = getIndexFromMap(column1)
    index2 = getIndexFromMap(column2)
    maxColumns = sheet.max_column
    newIndex = getIndexFromMap(newColumnName)
    if(newIndex == -1):
        sheet.cell(1,maxColumns+1).value = newColumnName
    columnArray[maxColumns] = newColumnName
    sumOfNewCell = 0
    for i in range(sheet.max_row-1) :
        if (i+1 < sheet.max_row) :
            value1 = sheet.cell(i+2,index1 + 1).value
            value2 = sheet.cell(i+2,index2 + 1).value
            if(newIndex == -1):
                sheet.cell(i+2,maxColumns+1).value =  value1 * value2
            sumOfNewCell = sumOfNewCell + value1 * value2
    wb.save(filePath)
    return sumOfNewCell

def stdDeviation(columnName) :
    index = getIndexFromMap(columnName)
    addColumnMinusMean(columnName)
    sumOfSquare = addSquareOfColumn(columnName+"MinusMean")
    return math.sqrt(sumOfSquare/(sheet.max_row-2))

def covariance(column1, column2) :
    addColumnMinusMean(column1)
    addColumnMinusMean(column2)
    sumOfMinusMean = addProductOfColumns(column1+"MinusMean" , column2+"MinusMean",column1+"MinusMean * "+column2+"MinusMean")
    return sumOfMinusMean/(sheet.max_row-2)

def correlation(column1, column2):
    correlationObj = Correlation()
    calCovariance = covariance(column1,column2)
    stdDeviation1 = stdDeviation(column1)
    stdDeviation2 = stdDeviation(column2)
    correlationObj.value = calCovariance / (stdDeviation1 * stdDeviation2)
    upperLimit = 1.96/math.sqrt(sheet.max_row -1 )
    if upperLimit < correlationObj.value :
        correlationObj.isSignificant = 1
    else :
        correlationObj.isSignificant = 0
    return correlationObj


def linearRegression(columnX, columnY):
    sumX = sumOfColumn(columnX);
    sumY = sumOfColumn(columnY);
    sumXY = addProductOfColumns(columnX, columnY, "sum"+columnX+columnY)
    sumXSq = addSquareOfColumn(columnX)
    linearReg = LinearRegression()
    print(sumXY)
    linearReg.m = (((sheet.max_row -1)*sumXY) - (sumY*sumX))/(((sheet.max_row -1 )*sumXSq) - (sumX*sumX))
    linearReg.c = (sumY - (linearReg.m *sumX))/(sheet.max_row - 1)
    return linearReg

if __name__ == "__main__" :
    main()
